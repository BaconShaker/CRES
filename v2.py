#!/usr/bin/python


# This is the second coming of the Jedi Order, v2.py.

# The idea here with v2 is basically mimic v1 in terms of program structure, 
# and change the file structure so instead of having one file to rule all other files, 
# we have one file and we just read the names to make the lists. 

# CRES
# 	-- Location 1 
# 		-- Detail 1
# 		-- Detail 2 
# 	-- Location 2 
# 		-- ..

# Also, you need to make sure the files that ARE NOT ON GITHUB (ie the location files)
# can all be found in '~/Locations'

# And in general clean things up a bit now that I know how Python Gooder. 


# So first, let us go and Import some libs and stuffs. 



import csv # This may need to be gotten with apt-get
import sys 
import operator # This is for what? Something in the arrange_locations() def
import os.path
from os.path import isfile, join
from datetime import datetime
import urllib
import urllib2
from bs4 import BeautifulSoup		# pip install BeautifulSoup4
import os
from os import listdir
from tabulate import tabulate # got this with 
import json # Pretty sure I got this for the GoogleMaps in dir2.py
from collections import defaultdict



# print "We are also going to need some packages..."
# for key in sys.modules():
# 	print key
# That will list the Imported Modules for you. 

# Now let's check which system is running.

# Need to find a way to get the GDrive file on the the Linux side of things. That's going to be hard


if sys.platform.startswith('darwin'):
	print ""
	print "This is not a Linux, it's a Mac so you're going to get lost on the keyboard."
	print "It would be cool if you could actually specify if it's Robby or Mike too\n"

	locfile = os.path.expanduser( "~/GDrive/cres_sheets" )  #mac
	print "Locfile: ", locfile, '\n'

elif sys.platform.startswith('linux'):
	print ""
	print 'This is a Linux\n' 
	locfile = os.path.expanduser( '~/cres_sheets2')
	print "Locfile: ", locfile, '\n'


# The first thing we're going to have to do is find the 
# file with all the locations in it. 
# Which should be easy if we put it in the right place.
# Then let's make a list of all the locations for the menu.

# locations = [location for location in locfile]
# print locations

# sheets is a list of files in locfile
sheets = [ f for f in listdir(locfile)   ] 


print "Unsorted: ", sheets, '\n'

# Use the .remove here to get rid of entries we don't want to make the list. 
sheets.remove('.DS_Store') # Not sure what this is but we don't want it.
sheets.remove('Mordor.xlsx') # This is the xlsx file that has all the csv files as worksheets, master is the first page. 
sheets.remove('master.csv') # A csv file that has all the addresses and contact information for each location. 

# Master should also have an updated 'last pickup' and 'next pickup' column in it. Still need to add that. 

print "Censored: ", sheets, '\n'
sheets.sort()
print "Sorted: ", sheets, "\n"


# Make sure we enter the main menu loop
menu_choice = [0, 'Main Menu']

# -------------------------------------------------------------------
# Given choices, what_to_do will print before, choices then after
# in a menu form and will return the [selection, words from the selection] 

def what_to_do(choices, before, after, default_choice, *args):
	# Returns a number corresponding to the response choices[index] in the list below. 
	os.system('clear') 	
	# choices.append('')
	looper = 0	
	while looper == 0:
		print "\n" , before
		print ""
		count = 0
		counter = [0,]
		for choice in choices:
			choice = choice.decode('utf-8')
			print "		" + str(count) + "	" + str(choice)
			counter.append(int(count))
			count = count + 1
		print ""
		print "** Default is set to [" + str(default_choice) + "]"
		print ""
		print after
		selection = raw_input()
		if selection == "":
			selection = default_choice
		else:
			selection = int(selection)
		if selection not in counter:
			looper = 0
			os.system('clear')
		else:
			print ""
			# choice = choice + 0   # I think this can be deleted
			word =  choices[int(selection)]
			print "You picked: [" + str(selection) + "]  " + str(word)
			# word = word.replace("\xe2\x80\x99" , "'")
			response = [ int(selection), word] 
			return response
			looper = looper + 1

# -----------------------------------------------------------------

class Client:
	def __init__(self, fmaster):
			opener = open(fmaster)
			all_lines = [ line.split(',') for line in opener ]
			self.place = all_lines[1]
			print '\n\n' , all_lines
			print ""

			opener.close()
		

	# Make a function that shows the header(s) of the csv's

	# Going to need to read the names of the CSV's then add it to the Restaurant Object
	# Also going to need to opn each file to get the actual information to add it to the Object

	def add(self):
		# Should probably check if the file exists already... I'll come back to this. 
		print "Place" , self.place
		pass

# -----------------------------------------------------------------

def make_locations(locfile):

	# Here is the list that controls the headers on the pickup_files
	pickup_heads = []

	# Open the master file and get the nicknames. The Nicknames will become the individual
	# filenames for the pickup files. 

	apple = open(locfile + '/master.csv')
	oranges = csv.DictReader(apple, dialect = 'excel', skipinitialspace = True)

	nicknames = [x['Name'] for x in oranges]

	for nick in nicknames:
		pickup_file = locfile + '/' + nick + '.csv'
		if os.path.exists(pickup_file):
			print "there's a file for: " , nick

		else:
			print "There is no '" ,  pickup_file   , "' we should make one!\n" 
			to_make = open( pickup_file, 'a' )
			writer = csv.DictWriter( to_make , fieldnames = pickup_heads)
			writer.writeheader()
			to_make.close() 
	return nicknames
	apple.close()

# -----------------------------------------------------------------

	

# -----------------------------------------------------------------

def class_loader(master_dir):
	# This method is USELESS! 

	projects = defaultdict(dict)
	
	handler = open(master_dir + '/master.csv')
	master_read = csv.DictReader( handler )
	
	headers = master_read.fieldnames
	for rowdict in master_read:
		if None in rowdict:
			del rowdict[None]
		name = rowdict.pop("Name")
		address = rowdict.pop('Street Address')
		projects[name][address] = rowdict

	return projects
	handler.close()

# -----------------------------------------------------------------

	# Looks up all the .csv files in cres_sheets, strips the extensions 
	# and compares the names to the choice made.

	# Need to make the individual .csv's have the same headers and get imputs
	# from pickups.py. 

def show_details(location, locfile):
	os.system('clear')
	
	fo = open(locfile + '/' + location + '.csv' )
	fr = csv.DictReader(fo, dialect = 'excel', skipinitialspace = True)

	# heads = fr.fieldnames
	# print fr.iteritems
	new_fr = []
	print ""
	count = 0
	to_show = [ 'pickup_count' , 'leftovers', 'score' , 'gallons_collected' , 'price' , 'income' , 'to_charity'  ]
	for row in fr:
		new_fr.append( { key.replace( '_' , ' ') : value for key, value in row.items() if key in to_show} )
 		count += 1
	# print '\n' * 10
	# print 'New fr: ' , new_fr
	print "These are the stats for" , location , ':\n\n' 
	# print "new_fr: " , new_fr
	print tabulate(new_fr, headers = 'keys') , '\n\n\n\n'

	print 'Press [ENTER] key to continue...\n'
	pause = raw_input()

	return tabulate(new_fr, headers = 'keys') , '\n\n\n\n'

	fo.close()


# -----------------------------------------------------------------

	# Should add the date --> blank = today

def run_pickup(spike):
	diesel = 'http://www.eia.gov/dnav/pet/pet_pri_gnd_dcus_r20_w.htm'
	ams = 'http://www.ams.usda.gov/mnreports/nw_ls442.txt'

	# File paths to look for:
	# Robby:
	pickups = "/Users/AsianCheddar/Desktop/Python/pickups.csv"
	locfile = "/Users/AsianCheddar/Desktop/Python/location_list.csv"
	# Mike:

	# Settings: 
	spacing = 2 # Gets used as an int later
	spacing = int(spacing)

	# Functions:

	def inches_cubed_to_gallons(inches3):
		inches3 = float(inches3) * 0.0043290
		return inches3

	def height_to_volume(height): 
		# Takes Height measured from bottom of tank to top of liquid and returns volume
		# dimensions of bin are: 36H x 28W x 48L
		h = 36
		w = 28
		l = 48
		volume = height * w * l
		gallons = inches_cubed_to_gallons(volume)
		return gallons

	def gallons_to_pounds(gallons):
		#weight of vegetable oil is 7.75lbs per gallon
		lbs = gallons * 7.75
		return lbs

	def price_lookup(pounds, ams):
		# Need to get work on this parser.  

		print ""
		print "Price_lookup():" # Input gallons or inches?
		print ""
		response = urllib2.urlopen(ams)
		# response = urllib2.urlopen('http://www.ams.usda.gov/mnreports/nw_ls442.txt')
		soup = BeautifulSoup(response)
		# page = response.readline()
		text = soup.get_text()
		print text[301:475]
		print ""
		# soup.prettify(formatter= 'html')
		# total = pounds * price
		# return total
		response.close()

	# GETS price from "diesel"
	def	get_a(grip):
		diesel = urllib2.urlopen(grip)
		soup = BeautifulSoup(diesel)
		# links = soup.find_all( "Current2")
		print "This is the price of Fuel today according to: ", diesel
		soup.prettify()
		data = soup.find_all('td' , 'Current2')
		length = len(data)
		# print data[13]
		temp = str(data[13])
		print temp
		price = temp[32:36]
		return price
		diesel.close()



	#---------------------------------------------------------------------------------------
	# First, clear the screen.

	os.system('clear')

	# Define locations for the menu
	sheets_two = [ sheet.replace( '.csv' , '' ) for sheet in sheets ]

	pickup_is_running = 1
	
	while pickup_is_running == 1:
		
	
		location_input = what_to_do(sheets_two, "Where would you like to run a pickup for?", 'Thank you!', 0)
		

		# Prompt for inputs (Loation, Height on Arrival, Height on Departure)
		inputs = {
			"location" : location_input[1]
		}

		print "\n" * spacing
		print "This is a pickup for", location_input[1] , ".\n"

		
		harrival = raw_input("				Height on Arrivial in [INCHES]  ")
		print "\n" * spacing

		steps = [ ['Height (arrivial): ' , harrival] ]
		status = tabulate(steps, headers = ['Step' , 'Value'])
		print status , '\n\n'

		
		hdepart = raw_input("				Height at Depature: [INCHES]  ")
		steps.append( [ 'Height (depart)' , hdepart ] )
		status = tabulate(steps, headers = ['Step' , 'Value'])
		print status
		print "\n\n"


		# I don't think this will work... yet, need to get hardware to make the inputs
		# just entering the time here will require decoding and too much time to be changed later anyway. 

		# print "What time did you arrive at ", location_input[1] , '?\n'
		# start_pickup_time = raw_input()

		# print "\n\nWhat time did you leave?"
		# end_pickup_time = raw_input()

		# So for now:
		print "		How long did you spend at " , location_input[1] , '?'
		pickup_duration = raw_input("							[HOURS]  ")
		steps.append( [ 'Time on location:' , pickup_duration ] )
		status = tabulate(steps, headers = ['Step' , 'Value'])
		print status

		print "\n\n"

		# Calculate volume -> pounds
		
		gal_arrival = height_to_volume(float(harrival))
		gal_departure = height_to_volume(float(hdepart))
		score = round(((float(gal_arrival) - float(gal_departure)) / float(gal_arrival)) ,2) * 100
		gallons_collected = gal_arrival - gal_departure
		pounds_collected = gallons_to_pounds(gallons_collected)


		# Calculate price using pounds

		# Look up AMS Price data
		price_lookup(pounds_collected, ams)

		print "Manually lookup the price and enter it here: [$cwt] Example, 23.34 \n\n"
		price = float(raw_input()) / 100.0
		flat_fee = 15.0 / 100.0 			# Here is where you change the flat fee
		print "Our flat fee is: " , flat_fee , "	It can be changed ~line 373"
		we_get = flat_fee * pounds_collected
		they_get = (price - flat_fee) * pounds_collected
		price_of_fuel = get_a(diesel) 

		# Start building inputs
		# Inputs is going to be the Dict that's RETURNed by this function. 
		inputs['height_on_arrival'] = harrival
		inputs['height_on_departure'] = hdepart
		inputs['score'] = score
		inputs['gallons_collected'] = round(gallons_collected, 2)
		inputs['pounds_collected'] = round(pounds_collected, 2)
		inputs['leftovers'] = round(gal_departure, 2)
		inputs['Pickup_Duration'] = pickup_duration
		inputs['income'] = round(we_get, 2)
		inputs['to_charity'] = round(they_get, 2) 
		inputs['price'] = price # of WVO
		inputs['fuel_price'] = price_of_fuel


	
		os.system('clear')
		# Calculate Fuel Surcharge
		# Datestamp
		# End Loop
		# Display resulting Dictionary from inputs[]
		print "Here is a breakdown of how the pickup went."
		print ""
		print "Location visited: %(location)s" % (inputs)
		print "Oil collected: %(gallons_collected)s gallons" % (inputs)
		print "We only left ~ %(leftovers)s gallons behind" % (inputs)
		print ""
		print "The price today was: %(price)s $cwt" % (inputs)
		print "Meaning CRES gets to keep, $%(income)s" % (inputs)
		print "And the Charities get $%(to_charity)s" % (inputs)
		print ""
		print "For a score of: %(score)s" % (inputs)
		print ""
		print "The price of #2 Diesel today: $%s" % (price_of_fuel)
		print "Source: " + diesel

		check = [ [ key.replace( '_' , ' ') , inputs[key] ] for key in inputs]
		print tabulate(check, headers = [ 'Input' , 'Value'])

		print "These results will be added to the pickups.csv file above."
		print "Thank you for your cooperation, we hope you come back soon!"
		print ""
		# for row in inputs:
		# 	print str(inputs[row]).replace('_' , ' ') , "	" , row 

		make_sure = raw_input("Would you like to continue?  ")
		if make_sure == "":
			pickup_is_running = 0
			return inputs
		else:
			pass
	
# -----------------------------------------------------------------

def add_to_csv(to_add):
	print "\n\nUSEAGE: add_to_csv( DICT to add )"
	print '\n\n\n\n'
	print to_add.keys()
	print ""
	print to_add
 	print '\n\n\n\n'
	# Figure out which location this info needs to be added to... 
	target_file = locfile + "/" + to_add['location'] + '.csv'


	print target_file
	print '\n\n\n\n'

	fop = open(target_file)
	fdr = csv.DictReader(fop, dialect = 'excel', skipinitialspace = True)


	pickup_count = 1

	for line in fdr:
		pickup_count += 1 
	print "This is pickup #:" , pickup_count

	to_add['pickup_count'] = pickup_count

	howlong_add = len(to_add.keys())
	howlong_fdr = len(fdr.fieldnames) 

	print "fdr: " , howlong_fdr
	print "to_add: " , howlong_add 

	if howlong_fdr == howlong_add:
		# This one only writes the INPUT ROW.
		fw = open(target_file, 'a')	
		writer = csv.DictWriter(fw, to_add.keys())
		writer.writerow(to_add)
		fw.close()

	else:
		# This overwrites the file completely, may be a good idea to make a script here that 
		# moves the file that's being replaced to a safe location?
		fw = open(target_file, 'wb')	
		writer = csv.DictWriter(fw, to_add.keys())
		writer.writeheader()
		writer.writerow(to_add)
		fw.close()

	fop.close()
	
# -----------------------------------------------------------------

def add_client(robby):
	os.system('clear')

	keep_adding = 0
	count = 0
	while keep_adding != "n":
		print "You have chosen to add a new location to the master file.\n"
		new_locations = {}

		print "What is the FULL name of the restaurant you would like to add?"
		new_locations["Name"] = raw_input("  ")
		print ""

		print "What is the NICKNAME you would like to assign this location?"
		new_locations['Nickname'] = raw_input("  ")
		print ""

		print "What is the street address of location to be added?"
		print "Eg: 2021 W Fulton"
		new_locations['Street Address'] = raw_input()
		print ""
		print "Make it able to go back to fix mistakes"
		print "What city is the location in? [Blank] = Chicago" 
		bb = raw_input() 
		if bb == "":
			new_locations['City'] = "Chicago"
		else: 
			new_locations['City'] = bb
		print ""

		print "What is the State? [BLANK] = IL"
		aa = raw_input() 
		if aa == "":
			new_locations['State'] = "IL"
		else: 
			new_locations['state'] = aa
		print ""

		print "What is the Zip?"
		new_locations['Zip'] = raw_input("	")
		print ""

		print "And lastly, who beith the Contactuth Personath?"
		new_locations['Contact_Person'] = raw_input("	")
		print ""

		print "Here is what you just entered, if it is correct hit [ENTER]"
		print ""
		new_row = []
		for key in new_locations:
			print key + ": " + new_locations[key]
			new_row.append(new_locations[key])

		print ""
		if raw_input() == "":
			locations = open(locfile  + '/master.csv', 'a')
			writer = csv.DictWriter(locations, new_locations.keys())
			writer.writerow(new_locations)
			locations.close()
			print ""

		print "Would you like to add another location right now? [Y/n]"
		print ""
		keep_adding = raw_input()
		
	

	return new_locations

# -----------------------------------------------------------------

def cap(s, l):

    return s if len(s)<=l else s[0:l-3]+'...'

# -----------------------------------------------------------------  

from openpyxl import Workbook
from openpyxl import worksheet
from openpyxl.workbook import Workbook
# Got these with pip install openpyxl

def write_to_xl(csvfiles):
	spartan = open(locfile + '/' + 'master.csv')
	lines = [ x.split(',') for x in spartan ]
	# Create the main Workbook
	master = Workbook()



	# Get to the first sheet, created by default when Workbook() is called
	main = master.active
	main.title = 'One To Rule Them All'
	for row in lines:
		main.append(row)
	

	# Now to create all the other sheets
	for restaurant in sheets:
		spam = open(locfile + '/' + restaurant)
		all_rows = [ line.split(',') for line in spam ]

		restaurant_name = restaurant.replace( ".csv" , "")
		restaurant_name = cap(restaurant_name, 31)

		ws = master.create_sheet()
		# Sometimes this adds a  header row for no reason. I think it has to do with whether or not the location was just created or not
		ws.header_footer.center_header.font_size = 14
		ws.header_footer.center_header.font_name = "Arial,Bold"

		ws.title = str(restaurant_name)
		for each_row in all_rows:
			ws.append( each_row )

	# Don't forget to save to .xlsx
	master.save(os.path.expanduser( "~/GDrive/cres_sheets/Mordor.xlsx" ))


# -----------------------------------------------------------------

import googlemaps

# This should be able to do all the mapping and directions. There's also a method that makes directions matrices
# Probably worth looking into at least for simplicity's sake.




# -----------------------------------------------------------------




main_menu = [
	'Main Menu', 
	'Add Client',
	'List Restaurants',
	'Run a pickup',
	'Write to xlsx',
	'EXIT',
	]	

while menu_choice[0] == main_menu.index('Main Menu'):
	# Let us start the Program here, methinks
	# Load the restaurants into the class. 
	name_list = make_locations(locfile)
	name_list = [name.decode('utf-8') for name in name_list]
	print "Name List: " , name_list

	# robby = Client(locfile + '/master.csv')
	# print "this is robby: " , robby.add()

	# robby = class_loader(locfile)
	# print robby['Erie Cafe']['536 W. Erie Street']



	main1 = "This is a list of what this program can currently do:"
	main2 = "Need to add: editing functionality for the master and pickup files during and after the script has run.\n"
	# menu_choice = what_to_do(main_menu, main1, main2, 0)
	# print "Menu Choice: " , menu_choice


	# re if pick below, 
	# 	[0] is the number input from the prompt,
	# 	[1] should be the string in 'word' from the list 
	default_choice = 0

	menu_choice = what_to_do(main_menu, main1, main2, default_choice)

	# menu_choice takes the int you type and compares the string to the list main_menu 
	if menu_choice[0] == main_menu.index('List Restaurants'):
		default_choice = 0
		top = "Here's a list of current clients. Choose a number to specific details."
		bottom = ""
		location_choice = what_to_do(name_list, top, bottom, default_choice)
		os.system('clear')
		show_details(location_choice[1], locfile)
		menu_choice = [0, 'Main Menu'] 


	elif menu_choice[0] == main_menu.index('Run a pickup'): # Run the pickup program, modify corresponding csv files and return to main_menu
		print '\n\nThis is going to run a pickup!'
		print '\n' * 10
		pickup_inputs = run_pickup('spike')
		add_to_csv(pickup_inputs)
		print '\n\n\n' * 3
		show_details(pickup_inputs['location'] , locfile )
		menu_choice = [0, 'Main Menu']

		
	elif menu_choice[0] == main_menu.index('Add Client'): # Add a location, return to main_menu
		print "This is where we should like to have the snake add a client. "
		add_client(321)
		menu_choice = [0,'Main Menu']


	elif menu_choice[0] == main_menu.index('EXIT'): # Quit the damn program already we don't want to do this anymore! 
		os.system('clear')
		break


	elif menu_choice[0] == main_menu.index('Write to xlsx'): # Write all the csv files to an xlsx file, each location has its own sheet. 
		write_to_xl(locfile)
		menu_choice = [0,'Main Menu']










	


